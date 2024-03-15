import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string
import os

class ExcelAutomator:
    def __init__(self,file_name):
        self.file_name = file_name
        print(file_name)
        self.month_and_extension = self.extract_month_and_extension()
        print(self.month_and_extension)
    
    def extract_month_and_extension(self):
        "extract month and extention from filename"
        return self.file_name.split('_')[1]
    
    def read_excel(self):
        "Read excel file"
        excel_df = pd.read_excel(self.file_name)
        return excel_df
    
    def create_pivot_table(self):
        "Create a pivot table from the excel data"
        excel_file = self.read_excel()
        return excel_file.pivot_table(index='Gender',columns='Product line', values='Total', aggfunc='sum').round(0)
    
    def save_report_to_excel(self, report_table, output_dir='/'):
        "Saves the pivot table from source to excel report"
        report_table.to_excel(f'report_{self.month_and_extension}', sheet_name="Report", startrow=4)
    
    def load_workbook(self):
        "Load workbook"
        return load_workbook(f'report_{self.month_and_extension}')
        
    def format_report(self, sheet):
        "Format report"
        month_name = self.month_and_extension.split('.')[0]
        sheet["A1"]= 'Sales Report'
        sheet["A2"]= month_name.title()
        sheet["A1"].font = Font('Arial', bold=True, size=20)
        sheet["A2"].font = Font('Arial',bold=True, size=20) 
    
    def apply_formula(self, sheet, min_column, max_column, min_row, max_row):
        "Applies formula to the sheet"
        alphabet =list(string.ascii_uppercase)
        excel_alphabet = alphabet[0:max_column]
        for i in excel_alphabet:
            if i !='A':
                sheet[f'{i}{max_row + 1}'] = f'=SUM({i}{min_row +1}:{i}{max_row})'
                sheet[f'{i}{max_row + 1}'].style = 'Currency'
            sheet[f'{excel_alphabet[0]}{max_row + 1}'] ='Total'
    
    def add_chart_to_sheet(self, sheet, min_column, max_column, min_row, max_row):
        "add bar column to the sheet"
        barChart = BarChart()
        data = Reference(sheet,min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row)
        categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row)
        barChart.add_data(data,titles_from_data=True)
        barChart.set_categories(categories)
        sheet.add_chart(barChart,"B12")
        barChart.title = 'Sales by Product Line'
        barChart.style = 2
    
    
    def automate_excel(self):
        "Automate the excel report generation"
        report_table = self.create_pivot_table()
        self.save_report_to_excel(report_table)
        wb = self.load_workbook()
        sheet = wb['Report']
        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row
        self.apply_formula(sheet,min_column,max_column,min_row,max_row)
        self.add_chart_to_sheet(sheet,min_column,max_column,min_row,max_row)
        self.format_report(sheet)
        wb.save(f'report_{self.month_and_extension}')
        print(report_table)
    
        
#memanggil class excelautomator
file_name = 'sales_2021.xlsx'
automator = ExcelAutomator(file_name)
automator.automate_excel()
